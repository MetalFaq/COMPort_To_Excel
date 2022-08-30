import sys
import os
import xlwings as xw
from openpyxl import load_workbook
from datetime import datetime
from dic_to_list import list_e_fase1, list_e_fase2, list_e_fase3,  list_phi_f1

# absoluta ubicación de este archivo .py
abs_path = os.path.dirname(__file__)

# Path donde se van a guardar las planillas cargadas con datos
relative_path_file = "/resultados"
full_path_file = abs_path + relative_path_file
if not os.path.exists(full_path_file):
    os.mkdir(full_path_file)

# Path Planilla SCVA
relative_path_template_scva = "/Planillas/Planilla_SCVA.xlsx"
full_path_template = abs_path + relative_path_template_scva

# Path Planilla p/normativa
relative_path_template_norma = "/Planillas/Planilla_Norma.xls"
full_path_template_norma = abs_path + relative_path_template_norma

def write_excel_scva(name_scva):
    name_file = full_path_file + '/SCVA_' + name_scva['num_serie'] + '.xlsx'

    if list_phi_f1[0] == '0': # Para identificar si es un archivo de activa.
        wb = load_workbook(filename=full_path_template)
        ws = wb.worksheets[0]
        ws['F4'] = datetime.now().date()
        ws['E14'] = name_scva['num_serie']
        ws['F14'] = list_e_fase1[0]
        ws['G14'] = list_e_fase1[1]
        ws['H14'] = list_e_fase1[2]
        ws['I14'] = list_e_fase1[3]
        ws['J14'] = list_e_fase1[4]
        ws['K14'] = list_e_fase1[5]                # F14, G14 y H14 carga alta fase 1. I14, J14 y K14 carga baja fase 1.

        if list_e_fase2:                                                       # Si pasa el If, es un medidor trifásico.
            ws['E56'] = name_scva['num_serie']
            ws['F56'] = list_e_fase2[0]
            ws['G56'] = list_e_fase2[1]
            ws['H56'] = list_e_fase2[2]
            ws['I56'] = list_e_fase2[3]
            ws['J56'] = list_e_fase2[4]
            ws['K56'] = list_e_fase2[5]

            ws['E98'] = name_scva['num_serie']
            ws['F98'] = list_e_fase3[0]
            ws['G98'] = list_e_fase3[1]
            ws['H98'] = list_e_fase3[2]
            ws['I98'] = list_e_fase3[3]
            ws['J98'] = list_e_fase3[4]
            ws['K98'] = list_e_fase3[5]
    else:                                   # por defecto escribe en la zona de reactiva. No verifica un PF especifico.
        try:
            wb = load_workbook(filename= name_file)
            ws = wb.worksheets[0]
            ws['L14'] = list_e_fase1[0]
            ws['M14'] = list_e_fase1[1]
            ws['N14'] = list_e_fase1[2]
            ws['O14'] = list_e_fase1[3]
            ws['P14'] = list_e_fase1[4]
            ws['Q14'] = list_e_fase1[5]

            ws['L56'] = list_e_fase2[0]
            ws['M56'] = list_e_fase2[1]
            ws['N56'] = list_e_fase2[2]
            ws['O56'] = list_e_fase2[3]
            ws['P56'] = list_e_fase2[4]
            ws['Q56'] = list_e_fase2[5]

            ws['L98'] = list_e_fase3[0]
            ws['M98'] = list_e_fase3[1]
            ws['N98'] = list_e_fase3[2]
            ws['O98'] = list_e_fase3[3]
            ws['P98'] = list_e_fase3[4]
            ws['Q98'] = list_e_fase3[5]
        except FileNotFoundError:
            print("Por favor, ingrese primero archivo de energia activa")
            sys.exit(1)

    wb.save(filename = name_file)

def write_excel_norma(data, type_num):
    # Problema! ---> Cuando el .txt cargado no tiene = longitud de datos especificada p/hacer el excel
    # no me genera el archivo.

    name_file = full_path_file + '/Planilla_Norma' + type_num['num_serie'] + '.xlsx'
    #name = 'G:/facundo/Escritorio/Herencia centurion/Parser_project_2/resultados/' + 'Planilla_Norma_' + type_num['num_serie'] + '.xlsx'

    if data[0]['phi'] == '0':  # Si es de activa, entra al if.
        wb = xw.Book(full_path_template_norma)
        sheet = wb.sheets['Hoja1']
        #Datos
        sheet.range('J6:N6').value = datetime.now().date()
        sheet.range('J9:K9').value = type_num['tipo']
        sheet.range('M9:N9').value = type_num['num_serie']
        #Valores de errores p/ cargas activas.
        sheet.range('I50:Q50').value = data[0]['error']
        sheet.range('I77:Q77').value = data[1]['error']
        sheet.range('I64:Q64').value = data[2]['error']
        sheet.range('I23:K23').value = data[3]['error']
        sheet.range('I36:K36').value = data[4]['error']
        sheet.range('L23:N23').value = data[5]['error']
        sheet.range('L36:N36').value = data[6]['error']
        sheet.range('O23:Q23').value = data[7]['error']
        sheet.range('O36:Q36').value = data[8]['error']
        sheet.range('I51:Q51').value = data[9]['error']
        sheet.range('I78:Q78').value = data[10]['error']
        sheet.range('I65:Q65').value = data[11]['error']
        sheet.range('I24:K24').value = data[12]['error']
        sheet.range('I37:K37').value = data[13]['error']
        sheet.range('L24:N24').value = data[14]['error']
        sheet.range('L37:N37').value = data[15]['error']
        sheet.range('O24:Q24').value = data[16]['error']
        sheet.range('O37:Q37').value = data[17]['error']
        sheet.range('I52:Q52').value = data[18]['error']
        sheet.range('I79:Q79').value = data[19]['error']
        sheet.range('I66:Q66').value = data[20]['error']
        sheet.range('I25:K25').value = data[21]['error']
        sheet.range('I38:K38').value = data[22]['error']
        sheet.range('L25:N25').value = data[23]['error']
        sheet.range('L38:N38').value = data[24]['error']
        sheet.range('O25:Q25').value = data[25]['error']
        sheet.range('O38:Q38').value = data[26]['error']
        sheet.range('I54:Q54').value = data[27]['error']
        sheet.range('I81:Q81').value = data[28]['error']
        sheet.range('I68:Q68').value = data[29]['error']
        sheet.range('I27:K27').value = data[30]['error']
        sheet.range('I40:K40').value = data[31]['error']
        sheet.range('L27:N27').value = data[32]['error']
        sheet.range('L40:N40').value = data[33]['error']
        sheet.range('O27:Q27').value = data[34]['error']
        sheet.range('O40:Q40').value = data[35]['error']
        sheet.range('I55:Q55').value = data[36]['error']
        sheet.range('I82:Q82').value = data[37]['error']
        sheet.range('I69:Q69').value = data[38]['error']
        sheet.range('I28:K28').value = data[39]['error']
        sheet.range('L28:N28').value = data[40]['error']
        sheet.range('O28:Q28').value = data[41]['error']
        sheet.range('I56:Q56').value = data[42]['error']
    else:
        try:
            wb = xw.Book(name_file)
            #filas de valores de error p/ carga reactiva
            sheet = wb.sheets['Hoja1']
            sheet.range('I111:Q111').value = data[0]['error']
            sheet.range('I131:Q131').value = data[1]['error']
            sheet.range('I122:Q122').value = data[2]['error']
            sheet.range('I90:K90').value = data[3]['error']
            sheet.range('I100:K100').value = data[4]['error']
            sheet.range('L90:N90').value = data[5]['error']
            sheet.range('L100:N100').value = data[6]['error']
            sheet.range('O90:Q90').value = data[7]['error']
            sheet.range('O100:Q100').value = data[8]['error']
            sheet.range('I112:Q112').value = data[9]['error']
            sheet.range('I132:Q132').value = data[10]['error']
            sheet.range('I123:Q123').value = data[11]['error']
            sheet.range('I91:K91').value = data[12]['error']
            sheet.range('I101:K101').value = data[13]['error']
            sheet.range('L91:N91').value = data[14]['error']
            sheet.range('L101:N101').value = data[15]['error']
            sheet.range('O91:Q91').value = data[16]['error']
            sheet.range('O101:Q101').value = data[17]['error']
            sheet.range('I113:Q113').value = data[18]['error']
            sheet.range('I133:Q133').value = data[19]['error']
            sheet.range('I124:Q124').value = data[20]['error']
            sheet.range('I92:K92').value = data[21]['error']
            sheet.range('I102:K102').value = data[22]['error']
            sheet.range('L92:N92').value = data[23]['error']
            sheet.range('L102:N102').value = data[24]['error']
            sheet.range('O92:Q92').value = data[25]['error']
            sheet.range('O102:Q102').value = data[26]['error']
            sheet.range('I115:Q115').value = data[27]['error']
            sheet.range('I135:Q135').value = data[28]['error']
            sheet.range('I126:Q126').value = data[29]['error']
            sheet.range('I94:K94').value = data[30]['error']
            sheet.range('I104:K104').value = data[31]['error']
            sheet.range('L94:N94').value = data[32]['error']
            sheet.range('L104:N104').value = data[33]['error']
            sheet.range('O94:Q94').value = data[34]['error']
            sheet.range('O104:Q104').value = data[35]['error']
            sheet.range('I116:Q116').value = data[36]['error']
            sheet.range('I136:Q136').value = data[37]['error']
            sheet.range('I95:K95').value = data[38]['error']
            sheet.range('L95:N95').value = data[39]['error']
            sheet.range('O95:Q95').value = data[40]['error']
            sheet.range('I117:Q117').value = data[41]['error']
        except FileNotFoundError:
            print('Por favor, ingrese primero archivo de energía activa ')
            sys.exit(1)

    wb.save(name_file)